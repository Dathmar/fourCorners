from django.db import models
from django.db.models.signals import post_save
from django.dispatch import receiver

from openpyxl import load_workbook

from django.db.models.functions import Concat


# Create your models here.


class AuctionHouse(models.Model):
    name = models.CharField(max_length=100)
    address = models.CharField(max_length=200)
    city = models.CharField(max_length=100)
    state = models.CharField(max_length=100)
    zip_code = models.CharField(max_length=10)
    phone_number = models.CharField(max_length=15)
    email = models.EmailField(blank=True, null=True)
    logo = models.ImageField(upload_to='logos')

    class Meta:
        verbose_name = 'Partner'
        verbose_name_plural = 'Partners'

    def __str__(self):
        return f'{self.name} | {self.city}'


class Auction(models.Model):
    auction_house = models.ForeignKey(AuctionHouse, on_delete=models.CASCADE, verbose_name='Partner')
    slug = models.SlugField(max_length=200)
    file = models.FileField(upload_to='auctions')

    __original_file = None

    class Meta:
        verbose_name = 'Sale'
        verbose_name_plural = 'Sales'

    def __str__(self):
        return f'{self.auction_house} | {self.slug}'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.__original_file = self.file

    def save(self, *args, **kwargs):
        process_excel = False
        if self.__original_file is None or self.__original_file != self.file:
            process_excel = True

        super().save(*args, **kwargs)

        if process_excel:
            self.process_excel()

    def process_excel(self):
        # delete previous items in the auction
        AuctionItem.objects.filter(auction=self).delete()

        wb = load_workbook(self.file.path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
            ai = AuctionItem.objects.create(
                auction=self,
                lot=f'Lot {row[0].value}',
                title=row[1].value,
                description=row[2].value,
                length=row[3].value if row[3].value != '' else 0,
                width=row[4].value if row[4].value != '' else 0,
                height=row[5].value if row[5].value != '' else 0,
            )
            ai.save()


class Box(models.Model):
    description = models.CharField(max_length=1000)
    max_item_length = models.DecimalField(max_digits=15, decimal_places=2)
    max_item_width = models.DecimalField(max_digits=15, decimal_places=2)
    max_item_height = models.DecimalField(max_digits=15, decimal_places=2)
    max_volume = models.DecimalField(max_digits=15, decimal_places=2, null=True, blank=True)

    class Meta:
        verbose_name = 'Box'
        verbose_name_plural = 'Boxes'

    def save(self, *args, **kwargs):
        if not self.max_volume:
            try:
                self.max_volume = self.max_item_height * self.max_item_length * self.max_item_width
            except ValueError:
                self.max_volume = 0
            except TypeError:
                self.max_volume = 0

        super(Box, self).save(*args, **kwargs)

    def __str__(self):
        return f'{self.description}'


class AuctionItem(models.Model):
    auction = models.ForeignKey(Auction, on_delete=models.CASCADE, verbose_name='Sale')
    lot = models.CharField(max_length=100)
    title = models.TextField(blank=True, null=True)
    tier = models.CharField(max_length=10, blank=True, null=True)
    description = models.TextField(blank=True, null=True)
    product_url = models.URLField(blank=True, null=True)
    photos_url = models.URLField(blank=True, null=True)

    length = models.DecimalField(max_digits=15, decimal_places=2, blank=True, null=True)
    width = models.DecimalField(max_digits=15, decimal_places=2, blank=True, null=True)
    height = models.DecimalField(max_digits=15, decimal_places=2, blank=True, null=True)
    volume = models.DecimalField(max_digits=15, decimal_places=2, blank=True, null=True)

    box = models.ForeignKey(Box, on_delete=models.CASCADE, blank=True, null=True)

    search_name = models.TextField(blank=True, null=True)

    def __str__(self):
        return f'{self.search_name}'

    def save(self, *args, **kwargs):
        if not self.search_name:
            self.search_name = f'{self.lot} | {self.title}'

        if not self.volume:
            try:
                self.volume = self.height * self.length * self.width
            except ValueError:
                self.volume = 0
            except TypeError:
                self.volume = 0

        if not self.box:
            self.box = self.get_packaging_for_item()

        super(AuctionItem, self).save(*args, **kwargs)

    def get_packaging_for_item(self):
        if self.volume == 0:
            return None

        boxes = Box.objects.filter(max_volume__gte=self.volume).order_by('max_volume')

        # loop through all the boxes to see if the item fits in it.
        for box in boxes:
            # we have to check each length, width, height combination to see if an item will fit in the box
            # I just do this manually, but we could have used permutations to do this.  I think this is faster because
            # the logic will stop once a box is found.
            if (self.length <= box.max_item_length and self.width <= box.max_item_width
                    and self.height <= box.max_item_height):
                return box

            if (self.length <= box.max_item_width and self.width <= box.max_item_height
                    and self.height <= box.max_item_length):
                return box

            if (self.length <= box.max_item_height and self.width <= box.max_item_length
                    and self.height <= box.max_item_width):
                return box

        return None

    class Meta:
        ordering = ['id']
        verbose_name = 'Asset'
        verbose_name_plural = 'Assets'

